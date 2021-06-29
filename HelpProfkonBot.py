from pprint import pprint
from tkinter import EXCEPTION

import pymysql
import traceback

from pymysql.cursors import DictCursor
from openpyxl import load_workbook, Workbook
from contextlib import closing


# wb = load_workbook('/Users/Diana/OneDrive/Python/DDkgta/GramotaBot/students.xlsx')
# sheets = wb.sheetnames
# for sheet in sheets:
#     print(sheet)

# sheet = wb.active
# print(sheet['A2'].value)
a = 0


def connectionDB():
    connect = pymysql.connect(
        host='IP',
        user='',
        password='',
        database='',
        cursorclass=DictCursor,)
    return connect

def isHE(section, table):
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'SELECT COUNT(*) as count FROM {table} WHERE section="{section}"'
                cursor.execute(query)
                if list(cursor)[0]['count'] > 0:
                    return True
                else:
                    return False
    except Exception as e:
        print('Ошибка: ', e) 


def get_count_doclad_inSection(section: str) -> dict:
    """
    Плоучает из базы данных doclad количество докладов по факультетам

    Args:
        section (str): Секция (Дисциплина)

    Returns:
        [dict]: {'ФАиЭ': int, 
                 'МТФ' : int, 
                 'ФЭиМ': int, 
                 'ЭМК' : int}
    """
    try: 
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'SELECT ФАиЭ, МТФ, ФЭиМ, ЭМК FROM doclad WHERE section = "{section}"'
                cursor.execute(query)
                cursor = list(cursor)
        return cursor[0]
    except Exception as e:
        print('Ошибка при получении количества докладов: ', e)

def get_count_mesto_inSection(section:str) -> dict:
    """[summary]
    Плоучает из базы данных количество занятых мест в дисциплине по факультетам\n 
    Args:
        section (str): Секция (Дисциплина)

    Returns:
        dict: {'1mesto': {'МТФ': 1, 'ФАиЭ': 0, 'ФЭиМ': 0, 'ЭМК': 0},
               '2mesto': {'МТФ': 1, 'ФАиЭ': 0, 'ФЭиМ': 0, 'ЭМК': 0},
               '3mesto': {'МТФ': 1, 'ФАиЭ': 0, 'ФЭиМ': 0, 'ЭМК': 0}}
    """

    try:
        top = dict()

        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                for mesto in range(3):

                    query = f'SELECT ФАиЭ, МТФ, ФЭиМ, ЭМК FROM {mesto+1}mesto WHERE section = "{section}"'
                    cursor.execute(query)
                    top[f'{mesto+1}mesto'] = list(cursor)[0]
                    # print(top)

        return top
    except Exception as e:
        print('Ошибка при получении количества мест: ', e)
        return top

def get_count_sections(isList: False) -> int or list:
    """[summary]
    Получает количество или название всех секций (Дисциплин)\n
    Args:
        isList (Bool): False - вернет количество всех секций (Дисциплин) (int)
                       True - вернет названия всех секций (Дисциплин) (list)

    Returns:
        int: Количество всех секций (Дисциплин)
         or
        list: Названия всех секций (Дисциплин)
    """
    try:    
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                if isList:
                    sections = []

                    query = f'SELECT section FROM doclad '
                    cursor.execute(query)
                    cursor = list(cursor)
                    for row in cursor:
                        sections.append(row['section'])
                    return sections

                else:
                    query = f'SELECT COUNT(*) FROM doclad '
                    cursor.execute(query)
                    cursor = list(cursor)
                    return cursor[0]['COUNT(*)']

    except Exception as e:
        print('Ошибка при получении количества секций: ', e)

def get_count_All_doclad() -> int:
    AllCountDoclad = 0
    try: 
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'SELECT SUM(ФАиЭ) as ФАиЭ, SUM(МТФ) as МТФ, SUM(ФЭиМ) as ФЭиМ , SUM(ЭМК) as ЭМК FROM doclad'
                cursor.execute(query)
                cursor = list(cursor)[0]
                for value in cursor.values():
                    AllCountDoclad += value

        return AllCountDoclad
    except Exception as e:
        print('Ошибка при получении количества секций: ', e)

def get_ves_ratio_section(section: str) -> float:
    """[summary]
    Получаем Весовой коэффициент секции\n
    Args:
        section (str): секций (Дисциплин)

    Returns:
        float: Весовой коэффициент секции
    """
    countDoclad1 = 0
    try:
        countAllDoclad = get_count_All_doclad()
        countAllSection = get_count_sections(isList=False)
        countDoclad = get_count_doclad_inSection(section=section)
        for value in countDoclad.values():
            countDoclad1 += value

        ratio = countDoclad1 * countAllSection / countAllDoclad

        return round(ratio,2)
    except Exception as e:
        print('Ошибка получения весовова коэффициента секции: ', e)

def get_sum_ratio_mesto_inFakyltet(mesto: int, fakyltet: str) -> float:
    
    if mesto == 1:
        rationMesto = 0.5
    elif mesto == 2:
        rationMesto = 0.3
    elif mesto == 3:
        rationMesto = 0.2
    
    try: 
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'SELECT SUM({fakyltet}) FROM {mesto}mesto_ratio'
                cursor.execute(query)
                cursor = list(cursor)[0]
                
        return round(cursor[f"SUM({fakyltet})"] + rationMesto, 2)
    except Exception as e:
        print('Ошибка при получении суммы коэффициентов места: ', e)

def get_victory_fakyltet() -> list:
    """[Получаем победителей ]

    Returns:
        list: 
            [('фэим', 1), ('эмк', 2), ('мтф', 3), ('фаиэ', 4)]
    """
    victors=[]
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'SELECT fakyltet, itogQualiti FROM helps ORDER BY itogQualiti ASC'
                cursor.execute(query)
                cursor = list(cursor)
                for cursor in cursor:
                    victors.append((cursor['fakyltet'], cursor['itogQualiti']))
                # print(victors[1][1])
                return victors
    except Exception as e:
        print('Ошибка получения победителей: ', e, traceback.format_exc())

# def get_quanlity_indicator_inFakyltet(fakyltet: str) -> float:
#     #TODO
#     try:  
#         with closing(connectionDB()) as connect:
#             with connect.cursor() as cursor:
#                 query = f'SELECT SUM({fakyltet}) FROM {mesto}mesto_ratio'
#                 cursor.execute(query)
#                 cursor = list(cursor)[0]
                
#         return round(cursor[f"SUM({fakyltet})"], 2)
#     except Exception as e:
#         print('Ошибка при получении суммы коэффициентов места: ', e)


def send_mesto_ration_all():
    # mesto=0 заполняет все места коэфициентами
    sections = get_count_sections(isList=True)
    fakyltetRation = {}
    try:  
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                
                for section in sections:
                    for indexMesto in range(3):
                        mestoInSection = get_count_mesto_inSection(section)[f'{indexMesto+1}mesto']
                        sumMesto = sum(mestoInSection.values())
                        vks = get_ves_ratio_section(section) 
                        print(mestoInSection, indexMesto+1)

                        if isHE(section=section, table=f"{indexMesto+1}mesto_ratio"):
                            for fakyltet in mestoInSection:
                                if mestoInSection[fakyltet] == 0:
                                    continue
                                query = f"""UPDATE {indexMesto+1}mesto_ratio SET {fakyltet} = {mestoInSection[fakyltet] * vks / sumMesto} WHERE section = '{section}'"""
                                cursor.execute(query)
                        else:
                            for fakyltet in mestoInSection:
                                if mestoInSection[fakyltet] == 0:
                                    continue
                                query = f"""INSERT INTO {indexMesto+1}mesto_ratio SET section = '{section}', {fakyltet} = {mestoInSection[fakyltet] * vks / sumMesto} """
                                cursor.execute(query)
                        connect.commit()
        # return round(cursor[f"SUM({fakyltet})"], 2)
    except Exception as e:
        print('Ошибка при отправке суммы коэффициентов места: ', e, traceback.print_exc())
        # connect.commit()

def send_mesto_quanlity_indicators():
    """
    Считает места по Качественному показателю
    """
    mesto = 0
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'SELECT fakyltet, qualityIndicator FROM helps ORDER BY `qualityIndicator` DESC'
                # DESC - сортировака от большего к меньшему
                # ASC - от меньшего к большему
                cursor.execute(query)
                rowsa = list(cursor)
                for row in rowsa:
                    mesto = mesto + 1
                    query1 = f"""UPDATE helps SET itogQualiti = {mesto} WHERE fakyltet = '{row["fakyltet"]}'"""
                    cursor.execute(query1)
                connect.commit()

    except Exception as e:
        print("Ошибка записи мест по качественному показателю: ", e)

def send_mesto_quantitative_indicators():
    """
    Считает места по Количественному показателю
    """
    mesto = 0
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'SELECT fakyltet, quantitativeIndicator FROM helps ORDER BY `quantitativeIndicator` DESC'
                cursor.execute(query)
                rowsa = list(cursor)
                for row in rowsa:
                    mesto = mesto + 1
                    query1 = f"""UPDATE helps SET itogQuantity = {mesto} WHERE fakyltet = '{row["fakyltet"]}'"""
                    cursor.execute(query1)
                connect.commit()

    except Exception as e:
        print("Ошибка записи мест по количественному показателю: ", e)

def send_ratio_section_inMesto_and_fakyltet(section: str, fakyltet: str, ratio: float, mesto: int):
    table = f"{mesto}mesto_ratio"
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                if isHE(section=section, table=table):
                    query = f'UPDATE {table} SET {fakyltet} = {ratio} WHERE section = "{section}"'
                    cursor.execute(query)
                    connect.commit()
                    return print("[OK]")

                query = f'INSERT {table} (section, {fakyltet}) VALUES (%s, %s)'
                cursor.execute(query,(section, ratio))
                connect.commit()
    except Exception as e:
        print("Ошибка записи весовова коэффициента секции: ", e)

def send_mesto_inSection(section: str, fakyltet: str, mesto: int):
    table = f"{mesto}mesto"
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:

                if isHE(section=section, table=table):
                    query = f'UPDATE {table} SET {fakyltet} = 1 WHERE section = "{section}"'
                    cursor.execute(query)
                    connect.commit()
                    return print("[OK]")

                query = f'INSERT {table} (section, {fakyltet}) VALUES (%s, %s)'
                cursor.execute(query,(section, 1))
                connect.commit()
    except Exception as e:
        print("Ошибка записи места: ", e)

def send_count_doclad(section: str, fakyltet: str, count: list, kafedra=None): 
    """[summary]
    Записывает в базу данных количество докладов факультета\n
    Если дисциплина уже есть то значение факультета обновиться\n 
    Если нет то создастся новая 

    Args:
        section (str): Секция (Дисциплина)
        fakyltet (str): Факультет
        count [list]: Количество докладов по факультетам
                      (ФАиЭ, МТФ, ФЭиМ, ЭМК)
        kafedra ([str], optional): Кафедра, по умолчанию нету
    """
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:

                if isHE(section, "doclad"):
                    query = f'UPDATE doclad SET {fakyltet} = {count} WHERE section = "{section}"'
                    cursor.execute(query)
                    connect.commit()
                    return print("[OK]")
                # print(count)
                # print(str(kafedra), str(section), str(count[0]), str(count[1]), str(count[2]), str(count[3] ))
                query = f'INSERT doclad (kafedra, section, ФАиЭ, МТФ, ФЭиМ, ЭМК) VALUES (%s, %s, %s, %s, %s, %s)'
                cursor.execute(query,(str(kafedra), str(section), str(count[0]), str(count[1]), str(count[2]), str(count[3] )))
                connect.commit()
    except Exception as e:
        print("Ошибка записи количества докладов: ", e, traceback.format_exc())

def send_people_info(fio: str, fakyltet: str, kafedra: str, groups: str, mesto: int, section: str, work: str, sex: str, ):
    
    try:
        with closing(connectionDB()) as connect:
            with connect.cursor() as cursor:
                query = f'INSERT people (fakyltet, kafedra, groups, fio, sex, mesto, section, nameWork) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)'
                cursor.execute(query,   (fakyltet, kafedra, groups, fio, sex, str(mesto), section, work))
                
                connect.commit()

    except Exception as e:
        print('Ошибка при записи человека: ', e, traceback.format_exc())
